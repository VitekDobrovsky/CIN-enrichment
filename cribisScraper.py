import openpyxl as opx
import pyautogui as pg
from time import sleep
import pyperclip

# name => ičo

class CribisScraper: 
    def __init__(self, path: str, nameCol: int, icoCol: int):
        self.path = path

        # sheet
        self.workbookObject = opx.load_workbook(path)
        self.sheet = self.workbookObject.active
        self.nameCol = nameCol
        self.icoCol = icoCol

        # cooldowns
        self.cooldowns = {
            "getIco iteration": 1,
            "company data loading": 2,
            "saving cooldown": 2
        }

        # positions
        self.positions = {
            "search bar": [764, 95],
            "1st result": [490, 365],
            "ičo start": [715, 415, 433], # 3rd item alternative y
            "ičo end": [790, 415, 433],   # 3rd item alternative y
        }

    def getName(self, col: int, row:  int) -> str: 
        return self.sheet.cell(row=row, column=col).value
    
    def getTotalRows(self, sheet) -> int:
        totalRows = 0
        for i in self.sheet:
            totalRows += 1
        
        return totalRows

    def getICOCribis(self, name: str) -> str:
        sleep(self.cooldowns["getIco iteration"])

        # copy full search url to clipboard
        url = "https://www3.cribis.cz/search/results?type=bs&q=" + name
        pyperclip.copy(url)
        print(name)
        
        # search name in Cribis
        pg.moveTo(self.positions["search bar"][0], self.positions["search bar"][1])
        pg.click()
        pg.hotkey("command", "v")
        pg.press("enter")
        # open compnay in Cribis
        pg.moveTo(1000,1000)
        pg.click()
        correct = input(":")
        if correct == "":
            pg.moveTo(self.positions["1st result"][0], self.positions["1st result"][1])
        elif correct == "p":
            ico = "x"
        else:
            sleep(2)

        sleep(1)

        pg.click()

        if correct != "p":
            pg.click()

            # copy IČO
            sleep(self.cooldowns["company data loading"])
            pg.moveTo(self.positions["ičo start"][0], self.positions["ičo start"][1])
            pg.mouseDown()
            pg.moveTo(self.positions["ičo end"][0], self.positions["ičo end"][1])
            pg.mouseUp()
            pyperclip.copy("x") # for no results found err
            pg.hotkey("command", "c")
            ico = pyperclip.paste()

            # if name of a company is too long that IČO jumps to 3rd line
            if ico == "Historie": 
                pg.moveTo(self.positions["ičo start"][0], self.positions["ičo start"][2])
                pg.mouseDown()
                pg.moveTo(self.positions["ičo end"][0], self.positions["ičo end"][2])
                pg.mouseUp()
                pg.hotkey("command", "c")
                ico = pyperclip.paste()
        
        return ico

    def run(self):
        # isReady = input("Did you loged into Cribis? (y/n)")
        # isReady = input("Do you have your browser and terminal window open that they dont overlap? (y/n)")
        countDown("click on your browser window! u have ", "s -_-", 5)
        totalRows = self.getTotalRows(self.sheet)
        # sheet 4loop
        row = 2
        for name in self.sheet:
            icoCell = self.sheet.cell(row=row, column= self.icoCol)

            #skip if ičo is already there
            if icoCell.value: 
                print(f"passed - cell {row} is already filled")
                row += 1
                continue 
            

            # company info
            name = self.getName(self.nameCol, row)
            ico = self.getICOCribis(name)

            # insert data
            icoCell.value = ico
            print(f"{name} => {ico}")

            # auto save
            if row % 5 == 0:
                self.workbookObject.save(self.path)
                print(f"saved! - currently on row {row} - {totalRows - row} remaining :)")


            row += 1
        

        self.workbookObject.save(self.path)

def countDown(preMsg: str, postMsg: str, amount: int):
    for i in range(amount, 0, -1):
        print(preMsg + str(i) + postMsg)
        sleep(1)


if __name__ == "__main__":
    CribisScraper("data/temp.xlsx", 1, 2).run()