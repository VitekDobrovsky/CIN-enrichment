import openpyxl as opx
import pyautogui as pg
from tools import *
from time import sleep


class Enricher: 
    def __init__(self, path: str):
        self.path = path

        # sheet
        self.workbookObject = opx.load_workbook(path)
        self.sheet = self.workbookObject.active
        self.nameCol = 1
        self.icoCol = 2

        # cooldowns
        self.cooldowns = {
            "getIco iteration": 1,
            "searching": 3.5,
            "company data loading": 3,
            "saving cooldown": 5
        }

        # positions
        self.positions = {
            "search bar": [764, 95],
            "1st result": [490, 365],
            "ičo start": [715, 415, 433], # 3rd item alternative y
            "ičo end": [765, 415, 433],   # 3rd item alternative y
        }

    def getName(self, col: int, row:  int) -> str: 
        return self.sheet.cell(row=row, column=col).value

    def getICO(self, name: str) -> str:
        sleep(self.cooldowns["getIco iteration"])

        # copy full search url to clipboard
        url = "https://www3.cribis.cz/search/results?type=bs&q=" + name
        copyToClipboard(url)
        
        # search name in Cribis
        pg.moveTo(self.positions["search bar"][0], self.positions["search bar"][1])
        pg.click()
        pg.hotkey("command", "v")
        pg.press("enter")
        sleep(self.cooldowns["searching"])

        # open compnay in Cribis
        pg.moveTo(self.positions["1st result"][0], self.positions["1st result"][1])
        pg.click()

        # copy IČO
        sleep(self.cooldowns["company data loading"])
        pg.moveTo(self.positions["ičo start"][0], self.positions["ičo start"][1])
        pg.mouseDown()
        pg.moveTo(self.positions["ičo end"][0], self.positions["ičo end"][1])
        pg.mouseUp()
        copyToClipboard("x") # for no results found err
        pg.hotkey("command", "c")
        ico = readFromClipboard()

        # if name of a company is too long that IČO jumps to 3rd line
        if ico == "Historie": 
            pg.moveTo(self.positions["ičo start"][0], self.positions["ičo start"][2])
            pg.mouseDown()
            pg.moveTo(self.positions["ičo end"][0], self.positions["ičo end"][2])
            pg.mouseUp()
            pg.hotkey("command", "c")
            ico = readFromClipboard()
        
        return ico

    def run(self):
        countDown("click on your browser window! you have ", "s -_-", 5)
    
        # sheet 4loop
        row = 2
        for name in self.sheet:
            icoCell = self.sheet.cell(row=row, column=self.icoCol)

            #skip if ičo is already there
            if icoCell.value: 
                print(f"passed - cell {row} is already filled")
                row += 1
                continue 
            

            # company info
            name = self.getName(self.nameCol, row)
            ico = self.getICO(name)

            # insert data
            icoCell.value = ico
            print(f"{name} => {ico}")

            # auto save
            if row % 5 == 0:
                self.workbookObject.save(self.path)
                print(f"saved! - currently on row {row} :)")
                if cooldown: countDown("cooling down :O ", "s", self.cooldowns["saving cooldown"])


            row += 1
        

        self.workbookObject.save(self.path)



if __name__ == "__main__":
    

    Enricher("data/temp.xlsx").run()