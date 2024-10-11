import requests
from bs4 import BeautifulSoup
import urllib.parse
import re
import openpyxl as opx


headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36"}

def getIco(name):
    query = f"{name} IČO"
    url = f"https://www.google.com/search?q={urllib.parse.quote(query)}"
    response = requests.get(url, headers=headers)
    searched_words = ["ičo", "IČO", "IČ", "ič"]


    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        results = soup.body.find_all(string=re.compile('.*{0}.*'.format("IČO")), recursive=True)


        for content in results:
                words = content.split()
                for i, word in enumerate(words):
                    if word not in ["dič", "DIČ"]:
                        if word in searched_words:
                            try:
                                ico = words[i + 1]
                                return ico
                            except:
                                continue


workbookObject = opx.load_workbook("./data/companies.xlsx")
sheet = workbookObject.active


for i, noneed in enumerate(sheet):
    if i == 0: continue

    name = sheet.cell(row=i, column=1).value
    try:
        ico = getIco(name)
        print(f"{name} ---> {ico}")
    except:
        print(f"{name} ---> no ico found")

    sheet.cell(row=i, column=2).value = str(ico)

    if i % 5 == 0:
        workbookObject.save("./data/ico.xlsx")
        print("saved")

workbookObject.save("./data/ico.xlsx")