import requests
import openpyxl as opx
from time import sleep


# ičo => legal name

def getName(ico: str):
        response = requests.get("https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty/" + str(ico))
        json = response.json()
        name = json["obchodniJmeno"]
        return name


def fixIco(ico: str):
    length = len(ico)
    reqLen = 8
    finalIco = None
    if reqLen - length <= 0:
        finalIco = ico
    else:
        for i in range(reqLen - length):
            pre = "0" * (i + 1)
            finalIco = pre + ico

    return finalIco


def getInfo(ico: str):
    response = requests.get("https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty/" + str(ico))
    json = response.json()
    info = {
        "name": json["obchodniJmeno"],
        "year": json["datumVzniku"],
        "street": json["adresaDorucovaci"]["radekAdresy1"],
        "city": json["sidlo"]["nazevObce"],
        "state": json["sidlo"]["nazevKraje"],
        "psc": json["sidlo"]["psc"]
    }
    
    return info

def fillInfo():
    workbookObject = opx.load_workbook("./data/ico.xlsx")
    sheet = workbookObject.active

    for i, row in enumerate(sheet):
        icoCell = sheet.cell(row=i + 1, column=2)
        legalNameCell = sheet.cell(row=i + 1, column=3)
        yearCell = sheet.cell(row=i + 1, column=4)
        streetCell = sheet.cell(row=i + 1, column=5)
        cityCell = sheet.cell(row=i + 1, column=6)
        stateCell = sheet.cell(row=i + 1, column=7)
        countryCell = sheet.cell(row=i + 1, column=8)
        pscCell = sheet.cell(row=i + 1, column=9)

        
        if not icoCell.value or i == 0:
            continue

        try:
            info = getInfo(icoCell.value)
            print(info)
        except:
            continue

        legalNameCell.value = info["name"]
        yearCell.value = info["year"]  
        streetCell.value = info["street"]
        cityCell.value = info["city"]
        stateCell.value = info["state"]
        pscCell.value = info["psc"]
        countryCell.value = "Česká republika"



        if i % 500 == 0:
            workbookObject.save("./data/ares.xlsx")


        sleep(0.3) 



    workbookObject.save("./data/ares.xlsx")



fillInfo()