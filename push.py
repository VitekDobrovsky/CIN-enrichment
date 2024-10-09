import pyautogui as pg
from support import countDown
from time import sleep
import pyperclip
import openpyxl as opx


positions = {
    "edit": [1625, 238],
    "ico": [1315, 368],
    "year": [1315,917],
    "rest": [1135, 917],
    "location": {
        "street": [1320, 713],
        "city": [1320, 775],
        "state": [1320, 835],
        "psc": [1320, 896],
        "country": [1320, 957]
    }
}
pg.click()


def push(name:str, ico:str, year:str = None, street:str = None, city:str = None, state:str = None, psc:str = None, country:str = None): # 5s

    # open company
    pyperclip.copy(name)
    pg.hotkey("command", "k")
    sleep(1)
    pg.hotkey("command", "v")
    sleep(2)
    pg.press("enter")
    pg.moveTo(positions["edit"][0], positions["edit"][1])
    sleep(1)
    pg.click()
    sleep(1)



    # add ico
    pyperclip.copy(str(ico))
    pg.moveTo(positions["ico"][0], positions["ico"][1])
    sleep(1)
    pg.click()
    pg.hotkey("command", "a")
    pg.hotkey("command", "v")

    # add year
    if year:
        try:
            pyperclip.copy(year)
            pg.moveTo(positions["year"][0], positions["year"][1])
            pg.click()
            pg.hotkey("command", "a")
            pg.hotkey("command", "v")
        except:
            pass

    pg.moveTo(positions["rest"][0], positions["rest"][1])
    pg.click()
    for i in range(20):
        pg.press("down")
    
    # add location
    if street:
        pyperclip.copy(street)
        pg.moveTo(positions["location"]["street"][0], positions["location"]["street"][1])
        pg.click()
        pg.hotkey("command", "a")
        sleep(0.2)
        pg.hotkey("command", "v")
    
    if city:
        pyperclip.copy(city)
        pg.moveTo(positions["location"]["city"][0], positions["location"]["city"][1])
        pg.click()
        pg.hotkey("command", "a")
        sleep(0.2)
        pg.hotkey("command", "v")
    
    if state:
        pyperclip.copy(state)
        pg.moveTo(positions["location"]["state"][0], positions["location"]["state"][1])
        pg.click()
        pg.hotkey("command", "a")
        sleep(0.2)
        pg.hotkey("command", "v")
    
    if psc:
        pyperclip.copy(psc)
        pg.moveTo(positions["location"]["psc"][0], positions["location"]["psc"][1])
        pg.click()
        pg.hotkey("command", "a")
        sleep(0.2)
        pg.hotkey("command", "v")
    
    if country:
        pyperclip.copy(country)
        pg.moveTo(positions["location"]["country"][0], positions["location"]["country"][1])
        pg.click()
        pg.hotkey("command", "a")
        sleep(0.2)
        pg.hotkey("command", "v")
    
    pg.press("enter")
    sleep(1)
    



workbookObject = opx.load_workbook("./data/final1.xlsx")
sheet = workbookObject.active




countDown("open bigns, u have ", "s", 3)


for i, row in enumerate(sheet):
    check = sheet.cell(row=i+1, column=10)
    if check.value == "pushed" or i == 0:
        continue


    push(sheet.cell(row=i+1, column=1).value, sheet.cell(row=i+1, column=2).value, sheet.cell(row=i+1, column=4).value, sheet.cell(row=i+1, column=5).value, sheet.cell(row=i+1, column=6).value, sheet.cell(row=i+1, column=7).value, sheet.cell(row=i+1, column=9).value ,sheet.cell(row=i+1, column=8).value)
    print(f"row {i+1} pushed")
    check.value = "pushed"

    if i % 10 == 0: workbookObject.save("./data/final1.xlsx")


