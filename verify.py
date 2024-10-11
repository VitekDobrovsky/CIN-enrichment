import openpyxl as opx
import pyperclip

workbookObject = opx.load_workbook("./data/verified.xlsx")
sheet = workbookObject.active


for rowNum, row in enumerate(sheet):
    if rowNum == 0: continue
    if sheet.cell(row=rowNum, column=10).value: continue

    rawName = sheet.cell(row=rowNum, column=1).value
    legalName = sheet.cell(row=rowNum, column=3).value
    kraj = sheet.cell(row=rowNum, column=7).value
    print(f"{rowNum}: {rawName}   |||   {legalName} -> {kraj}")

    while True:
        resp = input("-> ")
        if resp == "o":
            pyperclip.copy(rawName)
        elif resp == "k":
            pyperclip.copy(legalName)
        elif resp == "p":
            correct = True
            break
        else:
            correct = False
            break
    
    sheet.cell(row=rowNum, column=10).value = str(correct)

    if rowNum % 5 == 0:
        workbookObject.save("./data/verified.xlsx")
        print("saved")


workbookObject.save("./data/verified.xlsx")

        
    

